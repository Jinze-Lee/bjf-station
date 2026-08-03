# -*- coding: utf-8 -*-
"""
把台站原始数据预处理成符合 docs/数据接口规范.md 的 CSV

    原始数据/*.csv   （dataTaker 导出，含 TDP 温差）
    原始数据/*.xlsx  （径向变化仪导出，每 sheet 一个网关）
                    ↓  本脚本
    数据/<起>_<止>_观测数据.csv  +  _meta.json
                    ↓  工具/导入观测数据.py
    js/observations.js → 网页

用法：
    python 工具/预处理raw数据.py
    python 工具/预处理raw数据.py --src 原始数据 --out 数据
    python 工具/预处理raw数据.py --dry-run       # 只看统计，不写文件

===========================================================================
需要维护的只有下面这两张表：CHANNEL_MAP（通道→树号）和 SHEET_SKIP
新一批数据如果接线没变，直接跑就行；接线变了改 CHANNEL_MAP 即可。
===========================================================================
"""

import argparse
import csv
import io
import json
import re
import sys
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from pathlib import Path

BASE = Path(__file__).resolve().parent.parent
CST = timezone(timedelta(hours=8))

# ===========================================================================
# 配置 1：数采文件的 TDP 通道 → 树号
#
# 键 = 文件名里的样点标识（不区分大小写，出现在文件名中即匹配）
#   ZHL = 综合林样地   -> DT1
#   LYS = 落叶松样地   -> DT2
#   QXZ = 气象站下方   -> DT3
#
# 接线变了就改这里，不用动其它任何地方。
# ===========================================================================

# ===========================================================================
# 配置 0：从一堆导出文件里挑出该用的
#
# 数采平台每次导出都生成一个新文件，且**每个文件都包含从头到现在的全部记录**
# （比如 20260713ZHL.csv 里的数据是从 06-25 开始的）。所以同一时刻会在多个
# 文件里重复出现 —— 合并后必须去重，这个由 snap_series 负责（同格点只留
# 离格点最近的那个）。
#
# 规则（用户 2026-07-29 给定）：
#   保留  文件名 = 8 位日期 + ZHL / LYS / QXZ
#   排除  文件名里含 TEST 的（调试文件）—— 大小写不敏感，实测有 '20260727ZHL-test.csv'
# ===========================================================================
FILE_KEEP_RE = re.compile(r"^\d{8}(ZHL|LYS|QXZ)", re.I)
FILE_DROP_KW = "test"

CHANNEL_MAP = {
    # 每项写成 (树号, 中文树种名)。树种名不是注释 —— 脚本会拿它和 js/trees.js
    # 里该树号的 speciesZh 逐条比对，对不上直接报错终止。
    # 通道↔树号一旦接错，数据会安静地挂到别的树上，图表照画不误，光看结果发现不了。
    "ZHL": {                       # 综合林样地 -> DT1（用户 2026-07-28 核对）
        1: ("DT1-LDL1-1155", "辽东栎"),
        2: ("DT1-LDL2-1134", "辽东栎"),
        3: ("DT1-WJF1-1135", "五角枫"),
        4: ("DT1-BH1-1137",  "白桦"),
        5: ("DT1-HH1-1140",  "黑桦"),
        6: ("DT1-MD1-1149",  "蒙椴"),
        7: ("DT1-BH2-1132",  "白桦"),
    },
    "LYS": {                       # 落叶松样地 -> DT2（用户 2026-07-28 核对，与初版暂定顺序不同）
        1: ("DT2-WJF1-1141", "五角枫"),
        2: ("DT2-LYS1-1144", "落叶松"),
        3: ("DT2-MD1-1131",  "蒙椴"),
        4: ("DT2-LYS2-1151", "落叶松"),
        5: ("DT2-HTQ1-1248", "胡桃楸"),
        6: ("DT2-MD2-1249",  "蒙椴"),
        7: ("DT2-HTQ2-1157", "胡桃楸"),
    },
    "QXZ": {                       # 气象站下方样地 -> DT3（用户 2026-07-28 核对）
        1: ("DT3-LYS1-1150", "落叶松"),
        2: ("DT3-HH1-1142",  "黑桦"),
        3: ("DT3-SY1-1152",  "山杨"),
        4: ("DT3-SY2-1138",  "山杨"),
        5: ("DT3-HTQ1-1153", "胡桃楸"),
        6: ("DT3-HTQ2-1139", "胡桃楸"),
    },
}

# 配置 2：Excel 里不属于监测树木的 sheet（气象/土壤等），直接跳过
SHEET_SKIP = set()          # 留空 = 自动按 trees.js 的网关号筛，多余的 sheet 会被报告并跳过

# ---------------------------------------------------------------------------
# Granier 换算
#   K  = (ΔT₀ − ΔT) / ΔT
#   Fd = 0.0119 · K^1.231     [cm³ cm⁻² s⁻¹]  ×10000 → [g m⁻² s⁻¹]
# ΔT₀ 取法：**全期单一最大值**（用户选定）
# ---------------------------------------------------------------------------
GRANIER_A = 0.0119
GRANIER_B = 1.231
CM_S_TO_G_M2_S = 10000.0

# ---------------------------------------------------------------------------
# ΔT₀（零流量参考温差）怎么取 —— 这是本脚本对结果影响最大的一个选择
#
# "global"  全记录单一最大值。记录短时可用，**长记录下很危险**：
#           只要混进一个异常高的 ΔT（断电、加热丝异常、维护），ΔT₀ 就被抬上去，
#           而 K=(ΔT₀-ΔT)/ΔT 里 ΔT₀ 是分子的加数 —— **每一个点**的 Fd 都跟着系统性偏高。
#
#           实测（2026-07-29，记录从 5 天扩到 33 天）：
#             DT2-HTQ2-1157  ΔT₀ 比逐日最大值的中位数高 44.9%，
#                            平均日变化的峰/夜比从 36.9 掉到 2.9
#             DT2-LYS1/LYS2/WJF1 三棵树的全期最大值**都落在 07-15 22:30 同一时刻**
#                            —— 三棵树同时达到最大温差是设备事件，不是生理现象
#           而且多数 ΔT₀ 出现在 06-26~07-09，即**展示窗口之外**：
#           窗口外的一个坏点，污染了窗口内的每一个点。
#
# "moving"  以每个点为中心、前后各 DT0_WINDOW_DAYS/2 天窗口内的最大值（默认）。
#           容得下偶尔整夜有流量的日子，也跟得上传感器漂移与季节变化，
#           单个异常值只影响它附近几天而不是整条序列。这是长期序列的常规做法。
#
# "daily"   逐日最大值。最贴近"每天夜里都会归零"的假设；
#           若某夜确有夜间液流，当天 Fd 会被整体高估。
# ---------------------------------------------------------------------------
DT0_METHOD = "moving"
DT0_WINDOW_DAYS = 7

# ΔT 的合理量程；超出视为通道未接或故障，整条剔除
DT_MIN, DT_MAX = 0.05, 40.0

# ---------------------------------------------------------------------------
# 时间对齐
#
# 数采按整点/半点记录（13:00:00、13:30:00），径向变化仪是**独立设备**，
# 各自的时钟差几十秒到几分钟（实测 00:01:03、00:31:02）。
# 不对齐的话两条曲线时间戳永不相同，图表的 shared tooltip 只能显示其中一条。
#
# 用户确认「可以粗略匹配」，故统一吸附到最近的半点。
# 若两个原始点吸附到同一格，保留**离格点更近**的那个。
# ---------------------------------------------------------------------------
SNAP_MINUTES = 30

# ---------------------------------------------------------------------------
# 时间窗对齐
#
# 四个原始文件的记录窗口各不相同（数采三台各自开始导出的时间不同，
# 径向变化仪的 Excel 又是按固定 6 天窗口导出的）。不处理的话网页上
# 会出现「这段只有液流没有径向、那段反过来」的参差。
#
# 打开后：取各来源时间范围的**交集**，窗口外的点全部丢弃。
# 关掉：python 预处理raw数据.py --no-clip
#
# ⚠️ ΔT₀ 在**裁剪之前**用完整记录计算 —— 它是零流量参考基准，用的数据
#    越长越有机会捕捉到真正的零流量。若裁完再算，Fd 会被强制在窗口内
#    触底一次，那是方法的假象而非生理事实。
# ---------------------------------------------------------------------------
CLIP_TO_COMMON = True


def snap(dt):
    """吸附到最近的 SNAP_MINUTES 整倍数时刻"""
    step = SNAP_MINUTES * 60
    epoch = dt.timestamp()
    return datetime.fromtimestamp(round(epoch / step) * step, dt.tzinfo)


def snap_series(pts):
    """pts: [(dt, value)] -> 吸附后去重的 [(dt, value)]"""
    best = {}                       # 吸附时刻 -> (与格点的秒差, 原时刻, 值)
    for t, v in pts:
        s = snap(t)
        gap = abs((t - s).total_seconds())
        if s not in best or gap < best[s][0]:
            best[s] = (gap, t, v)
    return sorted((s, v) for s, (_, _, v) in best.items())


def load_trees():
    """从 js/trees.js 读取树号、网关号、中文树种名、所属样点"""
    src = (BASE / "js" / "trees.js").read_text(encoding="utf-8")
    rows = re.findall(
        r"id:\s*'([^']+)',\s*plot:\s*'([^']+)',.*?gateway:\s*(\d+),.*?speciesZh:\s*'([^']+)'",
        src, re.S)
    if not rows:
        raise SystemExit("!! 无法从 js/trees.js 解析树木信息，请先跑 工具/生成树木数据.py")
    ids = {r[0] for r in rows}
    gw2tree = {int(r[2]): r[0] for r in rows}
    species = {r[0]: r[3] for r in rows}
    plot = {r[0]: r[1] for r in rows}
    return ids, gw2tree, species, plot


def verify_channel_map(species, plot, log):
    """
    核对 CHANNEL_MAP 与 js/trees.js 是否一致。

    通道接错是这类数据最阴的错误：数据会安安静静地挂到别的树上，
    图表照画不误，曲线看着也正常，光看结果根本发现不了。
    所以这里做三重检查，任何一条不过就直接终止，不允许带着错误往下跑。
    """
    problems = []
    for site, mapping in CHANNEL_MAP.items():
        seen = {}
        for ch, item in sorted(mapping.items()):
            if not (isinstance(item, tuple) and len(item) == 2):
                problems.append("%s 通道 %d: 应写成 (树号, 中文树种名)" % (site, ch))
                continue
            tid, sp = item

            # ① 树号必须存在
            if tid not in species:
                problems.append("%s 通道 %d: 树号 %r 不在 js/trees.js 中" % (site, ch, tid))
                continue
            # ② 声明的树种必须与 trees.js 一致
            if species[tid] != sp:
                problems.append("%s 通道 %d: 树号 %s 在 trees.js 里是「%s」，这里写的是「%s」"
                                % (site, ch, tid, species[tid], sp))
            # ③ 同一棵树不能挂到两个通道
            if tid in seen:
                problems.append("%s: 树号 %s 同时出现在通道 %d 和 %d"
                                % (site, tid, seen[tid], ch))
            seen[tid] = ch

        # ④ 该样点的树是否全部用上（少了说明漏接，多了说明串了样点）
        tids = [t for t, _ in mapping.values() if isinstance(_, str)] or \
               [v[0] for v in mapping.values() if isinstance(v, tuple)]
        plots = {plot[t] for t in tids if t in plot}
        if len(plots) > 1:
            problems.append("%s: 映射跨了多个样点 %s" % (site, sorted(plots)))
        elif plots:
            pl = plots.pop()
            all_in_plot = {t for t, p in plot.items() if p == pl}
            missing = all_in_plot - set(tids)
            if missing:
                log.append("  [提示] %s(%s) 有 %d 棵树未接通道: %s"
                           % (site, pl, len(missing), ", ".join(sorted(missing))))

    if problems:
        print("!! CHANNEL_MAP 与 js/trees.js 不一致，已终止：")
        for x in problems:
            print("   - " + x)
        raise SystemExit(1)

    n = sum(len(m) for m in CHANNEL_MAP.values())
    log.append("  CHANNEL_MAP 校验通过：%d 个通道，树号与树种均与 trees.js 一致" % n)


def parse_dt(s):
    """解析各种时间戳写法。

    dataTaker 原样导出 : '2026/07/19 13:00:00.000'
    径向变化仪 Excel   : '2026-07-20 00:01:03'（openpyxl 也可能直接给 datetime）
    被 Excel 另存过的  : '2026/7/14 4:00 PM'   <- 12 小时制、无秒、月日不补零

    最后那种是坑：有人把 CSV 用 Excel 打开又保存，Excel 会按本机区域设置重写
    时间列。只认 24 小时制的话，这类文件会**一个点都解析不出来**却不报错，
    整份数据静悄悄地少掉。所以 12 小时制也必须认。
    """
    if isinstance(s, datetime):
        return s if s.tzinfo else s.replace(tzinfo=CST)
    s = str(s).strip()
    for fmt in ("%Y/%m/%d %H:%M:%S.%f", "%Y/%m/%d %H:%M:%S",
                "%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S",
                "%Y/%m/%d %I:%M:%S %p", "%Y/%m/%d %I:%M %p",
                "%Y-%m-%d %I:%M:%S %p", "%Y-%m-%d %I:%M %p",
                "%Y/%m/%d %H:%M", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(s, fmt).replace(tzinfo=CST)
        except ValueError:
            continue
    return None


# 只有时刻没有日期的时间戳，如 '3:00:00 PM' / '12:30:00 PM'。
# 同样是被 Excel 另存的后果，但这种**日期已经彻底丢了**，无法从文件内容还原。
# 猜日期等于伪造数据，所以只能整份跳过并明确报出来。
TIME_ONLY_RE = re.compile(r"^\d{1,2}:\d{2}(:\d{2}(\.\d+)?)?\s*([AaPp][Mm])?$")


# ===========================================================================
# 一、数采 CSV
# ===========================================================================
CH_RE = re.compile(r"(?:RRTDP30\s*\((\d+)\)|TDP30[_-](\d+))", re.I)


def read_datataker(path, site_key, valid_ids, log):
    """
    dataTaker 导出的一个文件里**交错着多种宽度的行**（多任务调度各写各的列组），
    只有与表头等宽的那批行才带 TDP 数据。这里按行宽筛，不能直接丢给 pandas。
    """
    rows = list(csv.reader(io.open(path, encoding="utf-8-sig", errors="replace")))
    if not rows:
        return {}
    hdr = rows[0]
    wide = [r for r in rows[1:] if len(r) == len(hdr)]

    # 先看时间列能不能用。被 Excel 另存过的文件可能只剩 '3:00:00 PM' 这种
    # 没有日期的时刻 —— 那样每一行都解析失败，但循环里只是 continue，
    # 结果是"读了个空文件"却毫无提示。这里提前判断并明确报错。
    probe = [r[0].strip() for r in wide[:20] if r and r[0].strip()]
    if probe and all(TIME_ONLY_RE.match(x) for x in probe):
        log.append("  [跳过] %s：时间戳只有时刻没有日期（如 %r）。"
                   % (path.name, probe[0]))
        log.append("         这是被 Excel 打开并另存过的后果，日期无法从文件内容还原，"
                   "猜日期等于伪造数据。")
        log.append("         请改用数采原始导出的那一份；若该时段别的文件已覆盖，可忽略。")
        return {"__SKIPPED__": path.name}

    log.append("  %s: 表头 %d 列，总 %d 行，其中含 TDP 的完整行 %d 行"
               % (path.name, len(hdr), len(rows) - 1, len(wide)))

    # 列号 -> 通道号
    ch_col = {}
    for i, c in enumerate(hdr):
        m = CH_RE.search(c)
        if m:
            ch_col[int(m.group(1) or m.group(2))] = i
    log.append("  %s: 识别到 TDP 通道 %s" % (path.name, sorted(ch_col)))

    mapping = CHANNEL_MAP.get(site_key, {})
    series = defaultdict(list)          # tree_id -> [(dt, ΔT)]
    skipped_ch, bad_pts = [], 0

    for ch in sorted(ch_col):
        if ch not in mapping:
            skipped_ch.append(ch)
            continue
        tid = mapping[ch][0]          # (树号, 中文树种名) —— 树种名由 verify_channel_map 校验
        col = ch_col[ch]
        for r in wide:
            dt = parse_dt(r[0])
            if dt is None:
                continue
            raw = r[col].strip()
            if raw in ("", "NaN", "nan"):
                continue
            try:
                v = float(raw)
            except ValueError:
                continue
            if not (DT_MIN <= v <= DT_MAX):     # 未接/故障通道的巨值在这里被拦掉
                bad_pts += 1
                continue
            series[tid].append((dt, v))

    if skipped_ch:
        log.append("  %s: 跳过未映射的通道 %s（CHANNEL_MAP 里没有）"
                   % (path.name, skipped_ch))
    if bad_pts:
        log.append("  %s: 剔除超出 ΔT 量程 [%g, %g] 的点 %d 个（未接或故障通道）"
                   % (path.name, DT_MIN, DT_MAX, bad_pts))
    return series


def dt0_series(pts):
    """给每个时刻算一个 ΔT₀。返回与 pts 等长的列表。

    见文件顶部 DT0_METHOD 的说明：长记录下用全期单一最大值会被异常值毒化。
    """
    if DT0_METHOD == "global":
        g = max(v for _, v in pts)
        return [g] * len(pts)

    if DT0_METHOD == "daily":
        byday = {}
        for t, v in pts:
            d = t.date()
            if v > byday.get(d, -1e9):
                byday[d] = v
        return [byday[t.date()] for t, _ in pts]

    # moving：以每点为中心、前后各半个窗口内的最大值。
    # pts 已按时间排序，用双指针滑窗，O(n) 而不是每点重扫一遍。
    half = timedelta(days=DT0_WINDOW_DAYS / 2.0)
    out, lo, hi = [], 0, 0
    n = len(pts)
    for i, (t, _) in enumerate(pts):
        while lo < n and pts[lo][0] < t - half:
            lo += 1
        while hi < n and pts[hi][0] <= t + half:
            hi += 1
        out.append(max(v for _, v in pts[lo:hi]) if hi > lo else pts[i][1])
    return out


def granier(series, log):
    """ΔT 序列 -> 液流通量密度 (g m^-2 s^-1)。"""
    out, table = {}, []
    for tid, pts in series.items():
        pts = snap_series(pts)
        d0s = dt0_series(pts)
        vals = []
        for (t, v), dt0 in zip(pts, d0s):
            k = (dt0 - v) / v if v > 0 else 0.0
            k = max(k, 0.0)                       # ΔT 略高于 ΔT₀ 时的数值噪声
            fd = GRANIER_A * (k ** GRANIER_B) * CM_S_TO_G_M2_S
            vals.append((t, fd))
        out[tid] = vals
        fds = [v for _, v in vals]
        table.append((tid, min(d0s), max(d0s), min(v for _, v in pts),
                      max(fds), sum(fds) / len(fds), len(vals)))

    log.append("")
    log.append("  ΔT₀ 取法: %s%s" % (DT0_METHOD,
               "（窗口 %g 天）" % DT0_WINDOW_DAYS if DT0_METHOD == "moving" else ""))
    log.append("  %-16s %8s %8s %8s %10s %10s %7s" %
               ("树号", "ΔT₀最小", "ΔT₀最大", "ΔT最小", "Fd最大", "Fd均值", "点数"))
    for tid, d0min, d0max, dtmin, fmax, fmean, n in sorted(table):
        log.append("  %-16s %8.3f %8.3f %8.3f %10.1f %10.1f %7d"
                   % (tid, d0min, d0max, dtmin, fmax, fmean, n))
    return out


# ===========================================================================
# 二、径向变化仪 Excel
# ===========================================================================
def read_dendro(path, gw2tree, log):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    series, skipped, nodiams = {}, [], []

    for name in wb.sheetnames:
        sheet_gw = None
        try:
            sheet_gw = int(str(name).strip())
        except ValueError:
            pass

        if sheet_gw is None or sheet_gw not in gw2tree or name in SHEET_SKIP:
            skipped.append(name)
            continue

        ws = wb[name]
        it = ws.iter_rows(values_only=True)
        try:
            hdr = [str(c).strip().lower() if c is not None else "" for c in next(it)]
        except StopIteration:
            continue
        if "ts" not in hdr or "diams" not in hdr:
            nodiams.append(name)
            continue
        i_ts, i_d = hdr.index("ts"), hdr.index("diams")

        pts = []
        for r in it:
            if r is None or len(r) <= max(i_ts, i_d):
                continue
            t = parse_dt(r[i_ts])
            if t is None or r[i_d] in (None, ""):
                continue
            try:
                mm = float(r[i_d])
            except (TypeError, ValueError):
                continue
            pts.append((t, mm * 1000.0))          # mm -> μm
        if pts:
            series[gw2tree[sheet_gw]] = snap_series(pts)

    wb.close()
    if skipped:
        log.append("  跳过与监测树木无关的 sheet: %s" % ", ".join(skipped))
    if nodiams:
        log.append("  跳过无 diams 列的 sheet: %s" % ", ".join(nodiams))

    log.append("")
    log.append("  %-16s %10s %10s %10s %7s" % ("树号", "读数最小", "读数最大", "全期变幅", "点数"))
    for tid in sorted(series):
        v = [x for _, x in series[tid]]
        log.append("  %-16s %10.1f %10.1f %10.1f %7d"
                   % (tid, min(v), max(v), max(v) - min(v), len(v)))
    return series


# ============================================================================
# 三、写出
# ===========================================================================
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--src", default=str(BASE / "原始数据"))
    ap.add_argument("--out", default=str(BASE / "数据"))
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--no-clip", action="store_true",
                    help="不裁剪到各来源的共同时间窗，保留各自完整范围")
    args = ap.parse_args()
    clip = CLIP_TO_COMMON and not args.no_clip

    src, out = Path(args.src), Path(args.out)
    if not src.exists():
        raise SystemExit("!! 原始数据目录不存在: %s" % src)
    out.mkdir(exist_ok=True)

    valid_ids, gw2tree, species, plot = load_trees()
    log = []
    verify_channel_map(species, plot, log)

    # --- 数采 CSV ---
    log.append("=" * 74)
    log.append("一、数采 CSV（TDP 温差 → 液流通量密度）")
    log.append("=" * 74)
    # --- 先按规则挑文件，再逐个读 ---
    # 子目录也要找：dendrometer 的 CSV 放在子文件夹里，正好会被 FILE_KEEP_RE 挡掉
    # （它们叫 20260701-LYS1.csv，日期后面多了个横杠，不匹配）。
    all_csv = sorted(src.rglob("*.csv"))
    picked, drop_name, drop_test = [], [], []
    for p in all_csv:
        if not FILE_KEEP_RE.match(p.stem):
            drop_name.append(p.name)
        elif FILE_DROP_KW in p.stem.lower():
            drop_test.append(p.name)
        else:
            picked.append(p)

    log.append("  扫描 %s" % src)
    log.append("  CSV 共 %d 个 -> 命名不符 %d 个，含 %s 的 %d 个，**保留 %d 个**"
               % (len(all_csv), len(drop_name), FILE_DROP_KW.upper(),
                  len(drop_test), len(picked)))
    if drop_test:
        log.append("    排除的调试文件: %s" % ", ".join(sorted(drop_test)))
    log.append("")

    raw_dt = {}
    sources = []       # 用于取交集：**按来源类型**聚合，不是按文件
    file_ranges = []   # 仅用于显示每个文件覆盖了什么
    skipped_files = []
    for p in picked:
        key = next((k for k in CHANNEL_MAP if k.lower() in p.stem.lower()), None)
        if key is None:
            log.append("  跳过 %s：文件名里没有 %s 任一标识"
                       % (p.name, "/".join(CHANNEL_MAP)))
            continue
        got = read_datataker(p, key, valid_ids, log)
        if "__SKIPPED__" in got:
            skipped_files.append(got["__SKIPPED__"])
            continue
        ts = [t for pts in got.values() for t, _ in pts]
        if ts:
            file_ranges.append((p.name, min(ts), max(ts)))
        for tid, pts in got.items():
            raw_dt.setdefault(tid, []).extend(pts)

    if skipped_files:
        log.append("")
        log.append("  [注意] 因时间戳损坏跳过 %d 个文件: %s"
                   % (len(skipped_files), ", ".join(skipped_files)))

    # 取交集必须按**来源类型**聚合，不能按文件。
    # 数采每次导出都是"从头到现在"的全量，35 个文件的时间范围互相嵌套；
    # 若按文件取 max(起点)，会被 20260728LYS.csv（起点 07-27 17:30）
    # 这种最新的小文件一把掐死，窗口塌成几小时。
    # 正确做法：先把所有液流文件**并**起来当作一个来源，再与径向生长取交集
    # —— 这也正是用户要的「把液流的时间汇总起来，再跟径向取子集」。
    if file_ranges:
        sources.append(("液流·合并 %d 个数采文件" % len(file_ranges),
                        min(r[1] for r in file_ranges),
                        max(r[2] for r in file_ranges)))

    fd = granier(raw_dt, log) if raw_dt else {}

    # --- 径向变化仪 Excel ---
    log.append("")
    log.append("=" * 74)
    log.append("二、径向变化仪 Excel（diams mm → 茎半径读数 μm）")
    log.append("=" * 74)
    # ⚠️ 多个 Excel 必须**合并**，不能 dendro.update(got)。
    #
    # dict.update 是整键替换：两个文件都有 1131 这棵树时，后读的会把先读的
    # 整条序列顶掉，而不是接起来。平台每次导出都是一个滚动窗口
    # （实测 (2)112 覆盖 07-10~07-28、(2)11 覆盖 07-27~07-31，谁也不含谁），
    # 用 update 的话前一段会**静默消失**，日志上还显示"读取成功"。
    #
    # 改成按树号 extend，再统一 snap_series 去重（同一格点保留离格点最近的）。
    dendro_raw = {}
    xlsx_ranges = []
    for p in sorted(src.glob("*.xlsx")):
        if p.name.startswith("~$"):
            continue
        log.append("  读取 %s" % p.name)
        got = read_dendro(p, gw2tree, log)
        ts = [t for pts in got.values() for t, _ in pts]
        if ts:
            xlsx_ranges.append((p.name, min(ts), max(ts)))
        for tid, pts in got.items():
            dendro_raw.setdefault(tid, []).extend(pts)

    dendro = {tid: snap_series(pts) for tid, pts in dendro_raw.items()}

    if xlsx_ranges:
        log.append("")
        log.append("  各 Excel 覆盖范围（互有重叠，合并后按格点去重）:")
        for nm, lo_, hi_ in sorted(xlsx_ranges, key=lambda r: r[1]):
            log.append("    %-40s %s ~ %s"
                       % (nm[:40], lo_.strftime("%m-%d %H:%M"), hi_.strftime("%m-%d %H:%M")))
        # 与液流一样：取交集时把所有 Excel 当作**一个**来源，否则会被
        # 覆盖最窄的那个文件掐死
        sources.append(("径向·合并 %d 个 Excel" % len(xlsx_ranges),
                        min(r[1] for r in xlsx_ranges),
                        max(r[2] for r in xlsx_ranges)))

    print("\n".join(log))

    if not fd and not dendro:
        print("\n没有解析出任何数据。")
        return 1

    # ---- 时间窗对齐：取各来源的交集 ------------------------------------
    n_before = sum(len(p) for p in fd.values()) + sum(len(p) for p in dendro.values())
    window = None
    print("\n" + "=" * 74)
    print("三、时间窗对齐")
    print("=" * 74)
    if file_ranges:
        print("  各数采文件覆盖范围（互相大量重叠，合并后按格点去重）:")
        for name, lo, hi in sorted(file_ranges, key=lambda r: r[1]):
            print("    %-26s %s ~ %s" % (name[:26], lo.strftime("%m-%d %H:%M"),
                                         hi.strftime("%m-%d %H:%M")))
        print()
    print("  参与取交集的来源:")
    for name, lo, hi in sources:
        print("    %-34s %s ~ %s" % (name[:34], lo.strftime("%Y-%m-%d %H:%M"),
                                     hi.strftime("%Y-%m-%d %H:%M")))

    if not clip:
        print("\n  --no-clip：不裁剪，保留各来源完整范围（网页上各曲线覆盖会参差）")
    elif len(sources) < 2:
        print("\n  只有一个来源，无需对齐")
    else:
        # 边界也要吸附到同一格点。否则会出现：某来源原始首点 15:00:00.006，
        # 吸附后落在 15:00:00，反而早于未吸附的窗口下界而被误裁掉。
        w_lo = snap(max(lo for _, lo, _ in sources))
        w_hi = snap(min(hi for _, _, hi in sources))
        if w_lo >= w_hi:
            print("\n  [错误] 各来源没有共同时间窗（最晚起点 %s 晚于最早终点 %s）"
                  % (w_lo.strftime("%m-%d %H:%M"), w_hi.strftime("%m-%d %H:%M")))
            print("  请检查是不是混入了不同批次的文件；用 --no-clip 可跳过对齐。")
            return 1
        window = (w_lo, w_hi)

        def clip_map(d):
            out = {}
            for tid, pts in d.items():
                kept = [(t, v) for t, v in pts if w_lo <= t <= w_hi]
                if kept:
                    out[tid] = kept
            return out

        fd, dendro = clip_map(fd), clip_map(dendro)
        n_after = sum(len(p) for p in fd.values()) + sum(len(p) for p in dendro.values())
        print("\n  共同窗口: %s ~ %s  (%.2f 天)"
              % (w_lo.strftime("%Y-%m-%d %H:%M"), w_hi.strftime("%Y-%m-%d %H:%M"),
                 (w_hi - w_lo).total_seconds() / 86400))
        print("  裁剪: %d -> %d 点，丢弃 %d 点 (%.0f%%)"
              % (n_before, n_after, n_before - n_after,
                 100 * (n_before - n_after) / max(n_before, 1)))
        print("  注: ΔT₀ 已在裁剪**之前**用完整记录算好，故窗口内 Fd 不一定触及 0")

    if not fd and not dendro:
        print("\n裁剪后没有剩余数据。")
        return 1

    # --- 汇总 ---
    all_t = ([t for pts in fd.values() for t, _ in pts] +
             [t for pts in dendro.values() for t, _ in pts])
    lo, hi = min(all_t), max(all_t)
    trees = sorted(set(fd) | set(dendro))

    print("\n" + "=" * 74)
    print("四、汇总")
    print("=" * 74)
    print("  时间范围  : %s ~ %s (UTC+8)"
          % (lo.strftime("%Y-%m-%d %H:%M"), hi.strftime("%Y-%m-%d %H:%M")))
    print("  树木数    : %d / %d" % (len(trees), len(valid_ids)))
    print("  液流点数  : %d  (%d 棵树)" % (sum(len(p) for p in fd.values()), len(fd)))
    print("  径向点数  : %d  (%d 棵树)" % (sum(len(p) for p in dendro.values()), len(dendro)))
    miss = valid_ids - set(trees)
    if miss:
        print("  [警告] 完全无数据的树: %s" % ", ".join(sorted(miss)))
    only_fd, only_dn = set(fd) - set(dendro), set(dendro) - set(fd)
    if only_fd:
        print("  [警告] 只有液流没有径向: %s" % ", ".join(sorted(only_fd)))
    if only_dn:
        print("  [警告] 只有径向没有液流: %s" % ", ".join(sorted(only_dn)))

    # 裁剪后各树各变量应当覆盖同一窗口 —— 这正是本次改动要达成的目标
    spans = {}
    for label, d in (("液流", fd), ("径向", dendro)):
        for tid, pts in d.items():
            spans[(label, tid)] = (pts[0][0], pts[-1][0])
    if spans:
        los = set(v[0] for v in spans.values())
        his = set(v[1] for v in spans.values())
        if len(los) == 1 and len(his) == 1:
            print("  覆盖一致  : 全部序列起止相同 ✓")
        else:
            print("  覆盖不一致: 起点 %d 种 / 终点 %d 种" % (len(los), len(his)))
            for k, v in sorted(spans.items()):
                if v[0] != min(los) or v[1] != max(his):
                    print("    %s %s: %s ~ %s" % (k[0], k[1],
                          v[0].strftime("%m-%d %H:%M"), v[1].strftime("%m-%d %H:%M")))

    if args.dry_run:
        print("\n--dry-run：未写出文件。")
        return 0

    # --- 写 CSV ---
    name = "%s_%s_观测数据" % (lo.strftime("%Y-%m-%d"), hi.strftime("%Y-%m-%d"))
    csv_path, meta_path = out / (name + ".csv"), out / (name.replace("观测数据", "meta") + ".json")

    rows = []
    for tid, pts in fd.items():
        for t, v in pts:
            rows.append((tid, t.isoformat(), "sap_flux_density", "%.2f" % v))
    for tid, pts in dendro.items():
        for t, v in pts:
            rows.append((tid, t.isoformat(), "stem_radius", "%.1f" % v))
    rows.sort(key=lambda r: (r[1], r[0], r[2]))

    with io.open(csv_path, "w", encoding="utf-8", newline="") as f:
        f.write("tree_id,timestamp,variable,value\n")
        for r in rows:
            f.write(",".join(r) + "\n")

    meta = {
        "batch": name,
        "produced_at": datetime.now(CST).isoformat(timespec="seconds"),
        "produced_by": "工具/预处理raw数据.py",
        "source": "dataTaker CSV (TDP ΔT) + 径向变化仪 Excel (diams)",
        "timezone": "Asia/Shanghai",
        "units": {"sap_flux_density": "g m-2 s-1", "stem_radius": "um"},
        "processing": {
            "granier": "K=(dT0-dT)/dT; Fd=%g*K^%g [cm3 cm-2 s-1] *%g -> g m-2 s-1"
                       % (GRANIER_A, GRANIER_B, CM_S_TO_G_M2_S),
            "dT0_method": ("moving-window maximum, %g days" % DT0_WINDOW_DAYS
                           if DT0_METHOD == "moving" else
                           "daily maximum" if DT0_METHOD == "daily" else
                           "whole-period maximum per channel"),
            "dT0_note": "记录延长到 33 天后，全期单一最大值被个别设备异常事件抬高"
                        "（实测某树 ΔT₀ 比逐日最大值中位数高 44.9%，另有三棵树的"
                        "全期最大值落在同一时刻），会使整条序列的 Fd 系统性偏高、"
                        "夜间基线抬升。故改用滑动窗口最大值。",
            "dT_range_filter": [DT_MIN, DT_MAX],
            "clip_to_common_window": bool(window),
            "common_window": [window[0].isoformat(), window[1].isoformat()] if window else None,
            "clip_note": "取四个原始文件时间范围的交集，窗口外的点全部丢弃，"
                         "使网页上所有曲线覆盖同一区间。ΔT₀ 在裁剪前用完整记录计算。",
            "time_snap_minutes": SNAP_MINUTES,
            "time_snap_note": "数采与径向变化仪是独立设备、时钟不同步（实测差几十秒到几分钟）。"
                              "统一吸附到最近的 %d 分钟格点，使两条曲线时间戳一致，"
                              "图表的 shared tooltip 才能并排显示两个变量。"
                              "同格冲突时保留离格点更近的原始点。" % SNAP_MINUTES,
            "stem_radius": "diams(mm) * 1000 -> um; 原始位移读数，零点由传感器安装位置决定，"
                           "图表按序列首点归零后显示为径向变化量",
        },
        "channel_map": CHANNEL_MAP,
        # notes 只在确有需要说明的情况时才写（如某台设备当期检修、某段数据剔除）。
        # 平常留空，页面上就不显示这一行。
    }
    meta_path.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")

    print("\n写出:")
    print("  %s  (%d 行, %.0f KB)" % (csv_path, len(rows), csv_path.stat().st_size / 1024))
    print("  %s" % meta_path)
    print("\n下一步:")
    print('  python "%s"' % (BASE / "工具" / "校验观测数据.py"))
    print('  python "%s"' % (BASE / "工具" / "导入观测数据.py"))
    return 0


if __name__ == "__main__":
    sys.exit(main())
