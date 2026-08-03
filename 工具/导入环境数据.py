# -*- coding: utf-8 -*-
"""
把气象站导出的环境数据整理成 js/environment.js

    桌面/data.xlsx  （气象站导出，每月一个 sheet）
            ↓  本脚本
    js/environment.js  →  网页上的「环境条件」图表

用法：
    python 工具/导入环境数据.py
    python 工具/导入环境数据.py --src "C:\\Users\\18256\\Desktop\\data.xlsx"
    python 工具/导入环境数据.py --dry-run

===========================================================================
两个必须留神的地方
---------------------------------------------------------------------------
① 各 sheet 的**列顺序不一样**（实测 202607 第 2 列是空气温度，202608 第 2 列
   是 PM2.5）。所以一律按**列名**取值，绝不能按位置。

② 时间窗要裁到与 js/observations.js 相同的区间。环境数据比树木数据两头都长
   （实测多出前 2 天、后 2 天），不裁的话两张图表的时间轴对不齐，
   联动拖动时会出现一张图已经到头、另一张还在滚的错位。
===========================================================================
"""

import argparse
import json
import re
import sys
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from pathlib import Path

BASE = Path(__file__).resolve().parent.parent
CST = timezone(timedelta(hours=8))
SNAP_MINUTES = 30

# ===========================================================================
# 变量表：Excel 列名 -> (键名, 单位, 是否默认显示, 画法)
#
# 排在前面的先显示在标签栏里。默认只开「光照 + 空气温度」两条 ——
# 它们是驱动蒸腾的直接因子，与上方液流曲线的对应关系最直观；
# 其余变量由用户按需点开。
#
# 画法：line 折线 / column 柱状 / scatter 散点
#   降雨用柱状：它是事件量不是连续量，画成折线会在两场雨之间连出斜线，
#              看着像"雨在慢慢变小"，其实中间根本没下。
#   风向用散点：0° 与 360° 是同一个方向，画折线会在越过正北时拉出一条
#              贯穿全图的竖线，纯属图形假象。
# ===========================================================================
VARS = [
    # Excel 列名          键名            单位          默认  画法
    ("光照klux",          "light",        "klux",       True,  "line"),
    ("空气温度℃",         "airTemp",      "°C",         True,  "line"),
    ("相对湿度%",         "rh",           "%",          False, "line"),
    ("降雨量mm",          "rain",         "mm",         False, "column"),
    ("土壤水分%",         "soilMoisture", "%",          False, "line"),
    ("土壤温度℃",         "soilTemp",     "°C",         False, "line"),
    ("气压hpa",           "pressure",     "hPa",        False, "line"),
    ("风速m/s",           "windSpeed",    "m/s",        False, "line"),
    ("风向°",             "windDir",      "°",          False, "scatter"),
    ("PM2.5ug/m³",       "pm25",         "µg/m³",      False, "line"),
    ("PM10ug/m³",        "pm10",         "µg/m³",      False, "line"),
]

# 不收的两列，各有各的理由：
#   电压V      设备状态，不是科学观测
#   径向生长mm 站点只有单点，而本站已有 20 棵树的逐树高精度径向记录，
#              放进来只会和树木数据打架
SKIP_COLS = {"电压V", "径向生长mm"}

# 合理量程，超出即视为故障或缺测哨兵值
RANGES = {
    "light": (0.0, 200.0), "airTemp": (-40.0, 60.0), "rh": (0.0, 100.0),
    "rain": (0.0, 500.0), "soilMoisture": (0.0, 100.0), "soilTemp": (-30.0, 60.0),
    "pressure": (500.0, 1100.0), "windSpeed": (0.0, 80.0), "windDir": (0.0, 360.0),
    "pm25": (0.0, 2000.0), "pm10": (0.0, 3000.0),
}


def parse_dt(s):
    if isinstance(s, datetime):
        return s if s.tzinfo else s.replace(tzinfo=CST)
    s = str(s).strip()
    for f in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S",
              "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M"):
        try:
            return datetime.strptime(s, f).replace(tzinfo=CST)
        except ValueError:
            pass
    return None


def snap(dt):
    step = SNAP_MINUTES * 60
    return datetime.fromtimestamp(round(dt.timestamp() / step) * step, dt.tzinfo)


def obs_window():
    """从 js/observations.js 读树木数据的时间窗，环境数据要裁到同一区间。

    OBS_META 是用 json.dumps(indent=2) 写出来的，所以键**带引号**、数组**跨行**：
        "span": [
          1752076800000,
          1753976...
        ],
    正则两点都要照顾到（\"?span\"? 和 \\s* 跨换行），否则匹配不上却又不报错，
    只会安静地跳过裁剪 —— 表现为两张图表时间轴对不齐。
    """
    p = BASE / "js" / "observations.js"
    if not p.exists():
        return None
    m = re.search(r'"?span"?\s*:\s*\[\s*(\d+)\s*,\s*(\d+)\s*\]',
                  p.read_text(encoding="utf-8"))
    if not m:
        return None
    return (datetime.fromtimestamp(int(m.group(1)) / 1000, CST),
            datetime.fromtimestamp(int(m.group(2)) / 1000, CST))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--src", default=r"C:\Users\18256\Desktop\data.xlsx")
    # 全量明文写到 原始数据/全量/ ——**不是** js/。
    # js/environment.js 现在是「最近 7 天」的公开版，由 工具/发布数据.py 切出来。
    # 往 js/ 写会把公开版覆盖成全量明文，然后随下一次 push 泄出去。
    ap.add_argument("--out",
                    default=str(BASE / "原始数据" / "全量" / "environment.js"))
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    src = Path(args.src)
    if not src.exists():
        raise SystemExit("!! 找不到环境数据文件: %s" % src)

    import openpyxl
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)

    col_of = {c: k for c, k, _, _, _ in VARS}
    merged = defaultdict(dict)      # 键名 -> {吸附时刻: 值}
    unknown, bad = set(), defaultdict(int)

    print("读取 %s" % src.name)
    for nm in wb.sheetnames:
        ws = wb[nm]
        it = ws.iter_rows(values_only=True)
        try:
            hdr = [str(c).strip() if c is not None else "" for c in next(it)]
        except StopIteration:
            continue
        if "时间" not in hdr:
            print("  sheet %r 没有「时间」列，跳过" % nm)
            continue
        i_ts = hdr.index("时间")

        for name in hdr:
            if name and name != "时间" and name not in col_of and name not in SKIP_COLS:
                unknown.add(name)

        n = 0
        for r in it:
            if not r or len(r) <= i_ts:
                continue
            d = parse_dt(r[i_ts])
            if d is None:
                continue
            d = snap(d)
            n += 1
            for i, name in enumerate(hdr):
                key = col_of.get(name)
                if key is None or len(r) <= i or r[i] in (None, ""):
                    continue
                try:
                    v = float(r[i])
                except (TypeError, ValueError):
                    continue
                lo, hi = RANGES.get(key, (-1e9, 1e9))
                if not (lo <= v <= hi):
                    bad[key] += 1
                    continue
                merged[key][d] = v
        print("  sheet %-8s %5d 行" % (nm, n))
    wb.close()

    if unknown:
        print("\n  [提示] Excel 里有本脚本未收录的列: %s" % ", ".join(sorted(unknown)))
        print("         要收进来的话，在 VARS 表里加一行即可。")
    for k, c in sorted(bad.items()):
        print("  [剔除] %s 超出量程的点 %d 个" % (k, c))

    if not merged:
        raise SystemExit("!! 没有解析出任何环境数据")

    # --- 裁到与树木数据相同的时间窗 ---
    win = obs_window()
    n_before = sum(len(v) for v in merged.values())
    if win:
        lo, hi = snap(win[0]), snap(win[1])
        for k in list(merged):
            merged[k] = {d: v for d, v in merged[k].items() if lo <= d <= hi}
        n_after = sum(len(v) for v in merged.values())
        print("\n  裁到树木数据的时间窗 %s ~ %s"
              % (lo.strftime("%Y-%m-%d %H:%M"), hi.strftime("%Y-%m-%d %H:%M")))
        print("  %d -> %d 点，窗口外丢弃 %d 点" % (n_before, n_after, n_before - n_after))
    else:
        print("\n  [警告] 读不到 js/observations.js 的时间窗，未做裁剪 ——")
        print("         两张图表的时间轴可能对不齐。")

    allt = sorted({d for v in merged.values() for d in v})
    if not allt:
        raise SystemExit("!! 裁剪后没有剩余数据（环境数据与树木数据的时间窗不重叠？）")

    print("\n  %-14s %10s %10s %10s %8s" % ("变量", "最小", "最大", "均值", "点数"))
    for col, key, unit, on, kind in VARS:
        if key not in merged or not merged[key]:
            print("  %-14s  (无数据)" % key)
            continue
        v = list(merged[key].values())
        print("  %-14s %10.2f %10.2f %10.2f %8d%s"
              % (key, min(v), max(v), sum(v) / len(v), len(v),
                 "   ← 默认显示" if on else ""))

    if args.dry_run:
        print("\n--dry-run：未写出文件。")
        return 0

    # --- 写 js/environment.js ---
    out = Path(args.out)
    L = []
    L.append("""/* 环境条件数据 —— 由 工具/导入环境数据.py 生成，请勿手工编辑。
 *
 * 来源：台站气象站导出（data.xlsx），30 分钟采样。
 * 时间窗已裁到与 js/observations.js 相同的区间，两张图表因此可以共用时间轴。
 *
 * ENV_VARS 每项：{ key, unit, on(默认是否显示), type(line/column/scatter) }
 * ENV_DATA[key] = [[毫秒时间戳, 值], ...]
 *
 * 这是**站点级**数据（单个气象站），不随选中的树木变化。
 */
""")
    meta = {
        "source": "station weather logger (data.xlsx)",
        "producedAt": datetime.now(CST).isoformat(timespec="seconds"),
        "timezoneOffsetMin": 480,
        "span": [int(allt[0].timestamp() * 1000), int(allt[-1].timestamp() * 1000)],
        "points": sum(len(v) for v in merged.values()),
    }
    L.append("var ENV_META = %s;\n" % json.dumps(meta, ensure_ascii=False, indent=2))

    L.append("var ENV_VARS = [")
    for col, key, unit, on, kind in VARS:
        if key in merged and merged[key]:
            L.append("    { key: %r, unit: %r, on: %s, type: %r },"
                     % (key, unit, "true" if on else "false", kind))
    L.append("];\n")

    L.append("var ENV_DATA = {")
    for col, key, unit, on, kind in VARS:
        if key not in merged or not merged[key]:
            continue
        pts = sorted(merged[key].items())
        # 小数位按量级定：气压要 1 位，PM 是整数，其余 2 位足够
        dec = 1 if key in ("pressure",) else (0 if key in ("pm25", "pm10", "windDir") else 2)
        body = ",".join("[%d,%.*f]" % (int(d.timestamp() * 1000), dec, v) for d, v in pts)
        L.append("  %r: [%s]," % (key, body))
    L.append("};\n")

    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text("\n".join(L), encoding="utf-8")
    print("\n写出: %s  (%.0f KB)" % (out, out.stat().st_size / 1024))
    print("  时间范围: %s ~ %s"
          % (allt[0].strftime("%Y-%m-%d %H:%M"), allt[-1].strftime("%Y-%m-%d %H:%M")))
    print("  变量 %d 个，合计 %d 点" % (len(ENV := [k for k in merged if merged[k]]), meta["points"]))
    print("\n这是**全量明文**，不进仓库。网页数据还没更新，还差一步：")
    print("      python 工具\\发布数据.py --pass \"你的口令\"")
    return 0


if __name__ == "__main__":
    sys.exit(main())
